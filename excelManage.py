#!/usr/bin/env python
# -*- coding: utf-8 -*-
import backend, random
import pandas as pd
from pandas import Series, DataFrame
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui
from showDataFrame import ShowDataFrame
from pandasModel import PandasModel

ASSES_CELL_COL    = "E"
LENGTH_CELL_COL   = "F"
ASSES_CELL_WIDTH  = 50
LENGTH_CELL_WIDTH = 20
GRADE_COL         = 0
CLASS_COL         = 1
STDNUM_COL        = 2
NAME_COL          = 3
ASSES_COL         = 4
LENGTH_COL        = 5

dataFrameList = []
classTextList = []

alignCenter        = Alignment(horizontal='center', vertical='center')
wrapText           = Alignment(vertical="center", wrapText=True)
allroundBorder     = Border(left =Side(border_style="thin",
                                        color='000000'),
                            right =Side(border_style="thin",
                                        color='000000'),
                            top   =Side(border_style="thin",
                                        color='000000'),
                            bottom=Side(border_style="thin",
                                        color='000000'))

#엑셀 테이블 항목 선택 시 내용 표시 함수
def exlActivateEdit(self):
    focusedItem = self.exlClassListWidget.currentItem()
    if(focusedItem is None):
        self.exlAseEdit.setPlainText("")
        return 
    if(focusedItem.column() == ASSES_COL):
        content = self.exlClassListWidget.currentItem().text()
        self.exlAseEdit.setPlainText(content)
        
def showAssesLengthAndState(self, currentContent, row):
    contentLength = len(currentContent)
    contentLengthByte = len(currentContent.encode("utf-8"))
    lengthItem = QTableWidgetItem(str(contentLength)+"자 ("+str(contentLengthByte)
                                    +"바이트)" )
    self.exlClassListWidget.setItem(row, LENGTH_COL, lengthItem)
    if(contentLength > 1000 or contentLengthByte > 3000):
        self.exlClassListWidget.item(row, LENGTH_COL).setBackground(QtGui.QColor(255,0,0))

def exlSaveToFile(self):
    global dataFrameList, classTextList
    dfList = dataFrameList
    clList = classTextList
    if(dfList == []): return QMessageBox.about(self, "주의", "엑셀로 저장할 항목들을 추가해주세요.")
    name, _ = QFileDialog.getSaveFileName(self, 'Save File','','Excel files (*.xlsx)')
    
    if(name == ""): return

    with pd.ExcelWriter(name, engine="openpyxl") as writer:
        for df, cl in zip(dfList, clList):
            df.to_excel(writer, sheet_name = cl, index=False)
            
        for cl in clList:
            worksheet = writer.sheets[cl]
            worksheet.column_dimensions[ASSES_CELL_COL].width = ASSES_CELL_WIDTH
            worksheet.column_dimensions[LENGTH_CELL_COL].width = LENGTH_CELL_WIDTH
            for col, i in zip(worksheet.columns, range(0, worksheet.max_column)):
                if i != ASSES_COL:
                    for cell in col:
                        cell.alignment = alignCenter
                        cell.border = allroundBorder
                if i == ASSES_COL:
                    for cell in col:
                        if cell.row == 1:
                            cell.alignment = alignCenter
                            cell.border = allroundBorder
                        else:
                            cell.alignment = wrapText
                            cell.border = allroundBorder
    dataFrameList = []
    classTExtList = []
        
def returnClassInteger(classes):
    classes = classes.replace(" ","")
    grade, ban = classes.split("학년")
    ban = ban.replace("반","")
    return int(grade), int(ban)

def exlAllShowClassList(self):
    widget = self.exlClassWidget
    widget.clear()
    classList = backend.returnClassList()

    for grade, classes in classList:
        item = QListWidgetItem(str(grade)+"학년 "+str(classes)+"반")
        widget.addItem(item)

def exlAddClassList(self):
    srcWidget = self.exlClassWidget
    targetWidget = self.exlAddedClassList
    if(srcWidget.selectedItems() is not None):
        items = srcWidget.selectedItems()
        for item in items:
            newItem = QListWidgetItem(item.text())
            targetWidget.addItem(newItem)

def exlExtClassList(self):
    widget = self.exlAddedClassList
    if(widget.currentItem() is not None):
        row = widget.currentRow()
        widget.takeItem(row)

def exlAllShowSubList(self):
    widget = self.exlSubWidget
    widget.clear()
    widget.setColumnCount(1)
    widget.setHeaderLabels(["과목"])
    parentSubjects = backend.returnParentSubject()
    for parent in parentSubjects:
        parentId = int(parent[0])
        parentName = parent[1]
        item = QTreeWidgetItem(widget, [parentName])
        item.setWhatsThis(0, str(parentId)+"-")
    
    it = QTreeWidgetItemIterator(widget)
    while it.value():
        if("-" in it.value().whatsThis(0)):
            parentId, trash = it.value().whatsThis(0).split("-")
            childSubjects = backend.returnChildSubjectsFromParentId(int(parentId))
            if(len(childSubjects) == 0):
                it += 1
            for child in childSubjects:
                childId = int(child[0])
                childName = child[1]
                item = QTreeWidgetItem(it.value())
                item.setWhatsThis(0, str(childId))
                item.setText(0, childName)
        it += 1

def exlAddSubList(self):
    srcWidget = self.exlSubWidget
    targetWidget = self.exlAddedSubWidget
    if(srcWidget.selectedItems() is not None):
        items = srcWidget.selectedItems()
        for item in items:
            if(item.parent() is not None): #자식 노드이면
                parentItem = item.parent()
                parentName = parentItem.text(0)
                childName = item.text(0)
                childId = item.whatsThis(0)
                newItem = QListWidgetItem(str(parentName) + " - " + str(childName))
                newItem.setWhatsThis(str(childId))
                targetWidget.addItem(newItem)

def exlExtSubList(self):
    widget = self.exlAddedSubWidget
    if(widget.currentItem() is not None):
        row = widget.currentRow()
        widget.takeItem(row)
                
def exlPrintMultiAsses(self):
    tabwidget = self.totalAssesTab
    tabwidget.clear()
    classList = []
    subjectIdList = []
    global dataFrameList
    global classTextList

    dataFrameList = []
    classTextList = []

    classListWidget = self.exlAddedClassList
    subjectListWidget = self.exlAddedSubWidget
    classListCnt = classListWidget.count()
    subjectListCnt = subjectListWidget.count()
    
    if(classListCnt == 0 or subjectListCnt == 0): return

    for i in range(0, classListCnt):
        classText = classListWidget.item(i).text()
        classTextList.append(classText)
        grade, classes = returnClassInteger(classText)
        classList.append([grade, classes])

    for i in range(0, subjectListCnt):
        subjectId = subjectListWidget.item(i).whatsThis()
        subjectIdList.append(int(subjectId))

    for item in classList:
        grade, classes = item
        gradeList = []
        classList = []
        assesment = []
        contentLengthList = []
        classMemberList = backend.returnClassMemberName(int(grade), int(classes))
        studentNumberList = backend.returnClassMemberNumber(int(grade), int(classes))
        studentIdList   = backend.returnClassMemberNumber(int(grade), int(classes))

        for i in range(0, len(studentNumberList)):
            num = str(studentNumberList[i])
            studentNumberList[i] = num[3:]

        for i in range(0, len(classMemberList)):
            gradeList.append(grade)
            classList.append(classes)
        rawData = {}
        rawData["학년"]  = gradeList
        rawData["반"]   = classList 
        rawData["번호"]  = studentNumberList
        rawData["이름"]  = classMemberList

        for studentId in studentIdList:
            
            assesText = ""
            tmpAsses = []
            for subjectId in subjectIdList:
                data = backend.returnStudentAssesmentBySubId(subjectId, studentId)
                if(data != []):
                    tmpAsses.append(data[0])

            #평가위치 셔플 
            if(self.assesmentShuffleCheckBox.isChecked()):
                random.shuffle(tmpAsses)
            
            for asses in tmpAsses:
                if(asses != ""):
                    asses.strip()
                    #줄바꿈모드 확인
                    if(self.lineChangeCheckBox.isChecked()):
                        assesText = assesText + " \n" + asses
                    else:
                        assesText = assesText + " " + asses

            assesText = assesText.strip()
            contentLength = len(assesText)
            contentLengthByte = len(assesText.encode("utf-8"))
            contentLengthList.append(str(contentLength) + " 자 (" + 
                                    str(contentLengthByte) + " 바이트)")
            assesment.append(assesText)
        
        rawData["평가"] = assesment
        rawData["글자수(바이트)"] = contentLengthList

        df = DataFrame(rawData)
        dataFrameList.append(df)

    for df, classText in zip(dataFrameList, classTextList):
        tab = QWidget()
        model = PandasModel(df)
        view = QTableView(tab)
        view.setModel(model)
        header = view.horizontalHeader()
        header.setSectionResizeMode(GRADE_COL, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(CLASS_COL, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(STDNUM_COL, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(NAME_COL, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(ASSES_COL, QHeaderView.Stretch)
        header.setSectionResizeMode(LENGTH_COL, QHeaderView.ResizeToContents)
        tabwidget.addTab(view, classText)
