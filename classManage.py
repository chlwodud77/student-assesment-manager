#!/usr/bin/env python
# -*- coding: utf-8 -*-
import backend
from openpyxl import load_workbook, Workbook
from PyQt5.QtWidgets import *
from PyQt5 import QtCore

def deleteStdClass(self):
    buttonReply = QMessageBox.question(self, '알림', "학급을 삭제하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    if buttonReply == QMessageBox.Yes:
        selectedItem = self.stdClassTreeWidget.currentItem()
        if(selectedItem is None):
            return QMessageBox.about(self, "주의", "삭제할 학급을 선택해주세요.")
        if("-" in selectedItem.whatsThis(0)):
            grade, classes = selectedItem.whatsThis(0).split("-")
            if(backend.deleteClass(int(grade), int(classes))):
                QMessageBox.about(self, "결과", "반 삭제 완료.")
                showClassList(self)  
            else:
                QMessageBox.about(self, "결과", "삭제 오류.")
                showClassList(self)
        else:
            return QMessageBox.about(self, "주의", "삭제할 학급을 선택해주세요.")

def deleteStd(self):
    buttonReply = QMessageBox.question(self, "알림", "선택 학생을 삭제하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    if (buttonReply == QMessageBox.Yes):
        selectedItem = self.stdClassTreeWidget.currentItem()
        if(selectedItem is None):
            return QMessageBox.about(self, "주의", "삭제할 학생을 선택해주세요.")
        if("-" in selectedItem.whatsThis(0)):
            return QMessageBox.about(self, "주의", "삭제할 학생을 선택해주세요.")
        stdId = selectedItem.whatsThis(0)
        if(backend.deleteStudent(int(stdId))):
            QMessageBox.about(self, "결과", "해당 학생 삭제 완료.")
            showClassList(self)
        else:
            QMessageBox.about(self, "결과", "삭제 오류.")
            showClassList(self)
        
def clsAddRow(self):
    ListWidget = self.stdListWidget
    ListWidget.insertRow(ListWidget.rowCount())
    
def clsDelRow(self):
    ListWidget = self.stdListWidget
    selectedRows = ListWidget.selectionModel().selectedRows()
    for r in sorted(selectedRows, reverse=True):
        ListWidget.removeRow(r.row())
        
def clsContentReset(self):
    stdListWidget = self.stdListWidget
    stdListWidget.clearContents()
    
def clsSetHeaders(self):
    headers = ["학번","이름"]
    widget = self.stdListWidget
    col = 2
    widget.setColumnCount(col)
    widget.setHorizontalHeaderLabels(headers)
    

def showClassList(self):
    classList = backend.returnClassList()
    classTreeWidget = self.stdClassTreeWidget
    classTreeWidget.clear()
    classTreeWidget.setColumnCount(1)
    classTreeWidget.setHeaderLabels(["학급"])
    for row in classList:
        parentItem = QTreeWidgetItem(classTreeWidget, [str(row[0])+"학년 "+str(row[1])+"반"])
        parentItem.setWhatsThis(0, str(row[0])+"-"+str(row[1]))
        
    it = QTreeWidgetItemIterator(self.stdClassTreeWidget)
    while it.value():
        if("-" in it.value().whatsThis(0)):
            grade, classes = it.value().whatsThis(0).split("-")
            students = backend.returnClassMemberList(int(grade), int(classes))
            for row in students:
                stdName = row[0]
                stdId = row[1]
                stdInfo = str(stdId)+" "+str(stdName)
                childItem = QTreeWidgetItem(it.value())
                childItem.setWhatsThis(0, str(stdId))
                childItem.setText(0, stdInfo)
        it += 1
        

def uploadCls(self):
    buttonReply = QMessageBox.question(self, '알림', "학급을 저장하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    if buttonReply == QMessageBox.Yes:
        if(self.stdListWidget.rowCount() == 0):
            return QMessageBox.about(self, "주의", "엑셀 파일로부터 학급 리스트를 불러와주세요.")
        stdList = self.stdListWidget
        rowCnt = stdList.rowCount()
        columnCnt = stdList.columnCount()

        # for col in range(0, columnCnt):
        #     header = self.stdListWidget.horizontalHeaderItem(col).text()
        #     grade = header[0]
        #     classes = header[2]
        #     for row in range(0, rowCnt):
        #         name = self.stdListWidget.item(row,col).text()
        #         if(name != "" or name is not None):
        #             backend.saveStudent(name, grade, classes)
        
        for row in range(0, rowCnt):
            if(stdList.item(row,0) is not None and stdList.item(row,1) is not None):
                hakbun = stdList.item(row, 0).text()
                name = stdList.item(row, 1).text()
                if(name != "" and hakbun != ""):
                    grade = hakbun[0]
                    classes = ""
                    if(int(hakbun[1]) == 0):
                        classes = hakbun[2]
                    else:
                        classes = hakbun[1]+hakbun[2]
                    if(not backend.saveStudent(int(hakbun), name, int(grade), int(classes))):
                        return QMessageBox.about(self, "결과", "학생 저장 실패. 이미 저장한 중복된 학급 또는 학생이 있나 확인하세요.")
        QMessageBox.about(self, "결과", "학생 저장 완료.")
        showClassList(self)
    
#학급 구성원 엑셀 파일로 불러와서 리스트로 보여줌
def uploadFile(self):
    fname = QFileDialog.getOpenFileName(self, 'Open file', './', 'Excel files (*.xlsx *.xls)')
    if(fname[0]):    
        wb = load_workbook(filename = fname[0]) 
        ws = wb["Sheet1"]
        max_row = ws.max_row
        max_col = 2

        widget = self.stdListWidget
        widget.setRowCount(max_row-1)
        widget.setColumnCount(max_col)
        headers = ["학번","이름"]
        widget.setHorizontalHeaderLabels(headers)

        for i in range(2, max_row+1):
            for j in range(1, max_col+1):
                cellValue = str(ws.cell(row=i, column= j).value)
                item = QTableWidgetItem(cellValue)
                widget.setItem(i-2,j-1,item)