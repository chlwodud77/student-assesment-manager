import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5 import QtCore
from openpyxl import load_workbook, Workbook
import sqlite3, random

#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("manager.ui")[0]


#화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.showSub(self.subTreeWidget)
        self.showSub(self.scoreSubTreeWidget)
        self.insertClassComboBox(self.classList)
        self.insertClassComboBox(self.exlClassList)
        self.exlShowClassList()

        #학급추가 탭
        self.fileUpdBtn.clicked.connect(self.uploadFile)
        self.clsSaveBtn.clicked.connect(self.uploadCls)

        #과목관리 탭
        self.addChildBtn.clicked.connect(self.addChildSub)
        self.subSaveBtn.clicked.connect(self.saveSub)
        self.subSrhBtn.clicked.connect(self.searchSub)
        self.subTreeWidget.itemClicked.connect(self.searchSub)
        self.subAddBtn.clicked.connect(self.addNewSubjectItem)
        self.subDelBtn.clicked.connect(self.delSub)
        self.grdAseDelBtn.clicked.connect(self.delAse)
        self.grdAseAddBtn.clicked.connect(self.addAse)
        self.grdAseModBtn.clicked.connect(self.modAse)
        self.grdAAseList.clicked.connect(self.activateEdit)
        self.grdBAseList.clicked.connect(self.activateEdit)
        self.grdCAseList.clicked.connect(self.activateEdit)

        #점수입력 탭
        self.scoreSubTreeWidget.itemDoubleClicked.connect(self.showSubClickedLabel)
        self.scoreSubTreeWidget.itemClicked.connect(self.showSubClickedLabel)
        self.saveAssesmentBtn.clicked.connect(self.saveAssesment)
        self.callClassMemberBtn.clicked.connect(self.showClassMemberList)
        self.createAssesmentBtn.clicked.connect(self.insertRandomAssesment)
        self.createIndiAssesmentBtn.clicked.connect(self.insertIndiRandomAssesment)

        #엑셀출력 탭
        self.exlClassList.activated.connect(self.exlShowClassList)
        self.exlSubAddBtn.clicked.connect(self.exlSubAddClass)
        self.exlSubExtBtn.clicked.connect(self.exlSubExtClass)
        self.printTotAssesBtn.clicked.connect(self.exlShowTotAssesment)
        self.exlFileSaveBtn.clicked.connect(self.exlSaveToFile)
    
    ##############엑셀출력###########################
    def returnClassAssesmentBySubId(self, subId, grade, classes):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT Student.name, Student.id, Score.asses FROM Student INNER JOIN Score ON Student.id = Score.stdId WHERE Score.subId = ? and Student.grade = ? and Student.class = ?"
                return conn.cursor().execute(sql, (subId, grade, classes)).fetchall()

        except sqlite3.IntegrityError:
            print("과목 조회 오류")

    def returnIfSubHasGrandParent(self, parentId):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT sub.id, sub.subName AS Sub, parentSub.subName AS parentSubName, parentSub.parentId AS grandParentId FROM Subject sub INNER JOIN Subject parentSub ON sub.parentId = parentSub.id WHERE sub.id = ?"
                return conn.cursor().execute(sql, (grade, classes)).fetchone()
        except sqlite3.IntegrityError:
            print("과목 조회 오류")

    def returnClassSubList(self, grade, classes):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "select DISTINCT Subject.id as subId, Subject.subName, Subject.parentId FROM Subject INNER JOIN Score ON Score.subId = Subject.id INNER JOIN Student ON Score.stdId = Student.id WHERE Student.grade = ? and Student.class = ?"
                return conn.cursor().execute(sql, (grade, classes)).fetchall()
        except sqlite3.IntegrityError:
            print("과목 조회 오류")

    #최종 엑셀 파일로 저장 함수
    def exlSaveToFile(self):
        filename = self.exlClassList.currentText()+".xlsx"
        wb = Workbook()
        ws = wb.active
        content = []
        for i in range(0, self.exlClassListWidget.rowCount()):
            p = []
            p.append(self.exlClassListWidget.item(i,0).text())
            if(self.exlClassListWidget.item(i,2) is not None):
                p.append(self.exlClassListWidget.item(i,2).text())
            else:
                p.append("")
            content.append(p)

        for x in range(1,len(content)+1):
            for y in range(1,3):
                ws.cell(row = x, column = y).value = content[x-1][y-1]

        wb.save(filename)
        QMessageBox.about(self, "결과", "저장 성공.")


        


    #학급별 종합 평가 출력 함수
    def exlShowTotAssesment(self):
        self.exlClassListWidget.clearContents()
        if(self.exlSubAddedWidget.count() == 0):
            QMessageBox.about(self, "오류", "과목을 추가해주세요.")
            return
        grade = int(self.exlClassList.currentText()[0])
        classes = int(self.exlClassList.currentText()[4])
        subjectsIds = []
        assesments = []
        members = []
        items = []
        for row in range(self.exlSubAddedWidget.count()):
            item = self.exlSubAddedWidget.item(row)
            items.append(item)
        for item in items:
            subjectsIds.append(int(item.whatsThis()))
        

        for id in subjectsIds:
            assesments.append(self.returnClassAssesmentBySubId(id, grade, classes))

        members = self.returnClassMemberList(grade, classes)
        self.exlClassListWidget.setRowCount(len(members))
        self.exlClassListWidget.setColumnCount(3)

        for i in range(0,len(members)):
            nameItem = QTableWidgetItem(str(members[i][0]))
            stdIdItem = QTableWidgetItem(str(members[i][1]))
            stdIdItem.setWhatsThis(str(members[i][1]))
            self.exlClassListWidget.setItem(i, 0, nameItem)
            self.exlClassListWidget.setItem(i, 1, stdIdItem)

        
        for i in range(0,len(members)):
            stdId = self.exlClassListWidget.item(i,1).text()
            for asses in assesments:
                for row in asses:
                    if(int(row[1]) == int(stdId)):
                        if(self.exlClassListWidget.item(i,2) is not None): #기존에 내용이 있는 경우 붙여서 추가
                            orgContent = self.exlClassListWidget.item(i,2).text()
                            newContent = orgContent+" "+str(row[2])
                            assesItem = QTableWidgetItem(newContent)
                            self.exlClassListWidget.setItem(i,2, assesItem)
                        else: #기존에 내용이 없는 빈 칸인 경우
                            assesItem = QTableWidgetItem(str(row[2]))
                            self.exlClassListWidget.setItem(i,2, assesItem)

    #학급별 평가 과목 선택 추가 함수
    def exlSubAddClass(self):
        if(self.exlSubListWidget.currentItem() is not None):
            item = self.exlSubListWidget.currentItem()
            addItem = QListWidgetItem(str(item.text()))
            addItem.setWhatsThis(str(item.whatsThis()))
            self.exlSubAddedWidget.addItem(addItem)


    #학급별 평가 과목 선택 빼기 함수
    def exlSubExtClass(self):
        if(self.exlSubAddedWidget.currentItem() is not None):
            row = self.exlSubAddedWidget.currentRow()
            self.exlSubAddedWidget.takeItem(row)



    #학급별 평가 과목 보여주는 함수
    def exlShowClassList(self):
        self.exlSubListWidget.clear()
        self.exlSubAddedWidget.clear()
        grade = int(self.exlClassList.currentText()[0])
        classes = int(self.exlClassList.currentText()[4])
        subjects = self.returnClassSubList(grade,classes)
        subjects = sorted(subjects)

        for subject in subjects:
            parentSubName = self.returnSubNameBySubId(int(subject[2]))
            name = parentSubName+" - "+subject[1]
            item = QListWidgetItem(str(name))
            item.setWhatsThis(str(subject[0]))
            self.exlSubListWidget.addItem(item)


    ##############점수입력###########################

    def returnSubIdBySubName(self, subName):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT id FROM Subject WHERE subName = ?"
                return conn.cursor().execute(sql, (subName,)).fetchone()[0]
        except sqlite3.IntegrityError:
            print("과목 조회 오류")
    
    def returnSubNameBySubId(self, subId):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT subName FROM Subject WHERE id = ?"
                return conn.cursor().execute(sql, (subId,)).fetchone()[0]
        except sqlite3.IntegrityError:
            print("과목 조회 오류")

    def returnAssesmentContentById(self, id):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT content FROM Assesment WHERE id = ?"
                return conn.cursor().execute(sql, (id,)).fetchone()[0]
        except sqlite3.IntegrityError:
            print("과목 평가문 조회 오류")

    def returnAssementStandardBySubId(self, subId):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT DISTINCT grade, greater, less FROM Assesment WHERE subId = ?"
                return conn.cursor().execute(sql, (subId,)).fetchall()
        except sqlite3.IntegrityError:
            print("과목 조회 오류")

    def returnAssesmentBySubId(self, subId):                         
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT id, grade, content FROM Assesment WHERE subId = ?"
                return conn.cursor().execute(sql, (subId,)).fetchall()
        except sqlite3.IntegrityError:
            print("과목 평가문 조회 오류")


    def returnClassMemberList(self, grade, classes):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT name, id FROM Student WHERE grade = ? and class = ?"
                return conn.cursor().execute(sql, (grade, classes)).fetchall()
        except sqlite3.IntegrityError:
            print("학급 구성원 조회 오류")

    def returnClassList(self):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "SELECT DISTINCT grade, class from Student"
                return conn.cursor().execute(sql).fetchall()
        except sqlite3.IntegrityError:
            print("학급 조회 오류")

    def returnSubNameList(self):
        conn = sqlite3.connect("studentManager.db")
        # conn.row_factory = lambda cursor, row: row[0]
        try:
            with conn:
                sql = "SELECT a.Id, a.subName as childSub,  b.subName as parentSub FROM Subject a INNER JOIN Subject b ON a.parentId = b.id"
                return conn.cursor().execute(sql).fetchall()
        except sqlite3.IntegrityError:
            print("과목 조회 오류")
    
    def deleteScoreById(self, id):
        conn = sqlite3.connect("studentManager.db")
        try:
            with conn:
                sql = "DELETE FROM Score WHERE id = ?"
                conn.cursor().execute(sql, (id,))
        except sqlite3.IntegrityError:
            print("점수 삭제 오류")


    #과목 선택 시 라벨에 선택 과목 표시 함수
    def showSubClickedLabel(self):
        self.subClickedLabel.setText(self.scoreSubTreeWidget.currentItem().text(0))

    def saveScore(self, params):
        conn = sqlite3.connect("studentManager.db")
        sql = "INSERT into Score(subId, stdId, score, asses) VALUES (?,?,?,?)"
        try:
            with conn:
                c = conn.cursor()
                c.execute(sql, (params[0], params[1], params[2], params[3]))
                return True
        except sqlite3.IntegrityError:
            print("점수 저장 오류")
            QMessageBox.about(self, "결과", "점수 저장 실패.")
            return False

    #점수로 생성된 평가문 저장해주는 함수
    def saveAssesment(self):
        if(self.scoreSubTreeWidget.currentItem() is None):
            QMessageBox.about(self, "오류", "과목을 선택해주세요.")
            return
        if(self.scoreSubTreeWidget.currentItem() is not None and self.classListWidget.rowCount() == 0):
            QMessageBox.about(self, "오류", "학급을 불러와주세요.")
            return
        for row in range(0, self.classListWidget.rowCount()):
            if(self.classListWidget.item(row,2).whatsThis() != ""): #기존 스코어 존재
                # assesId = int(self.classListWidget.item(row,3).whatsThis())
                asses = self.classListWidget.item(row,3).text()
                score = int(self.classListWidget.item(row,2).text())
                scoreId = int(self.classListWidget.item(row,2).whatsThis())
                stdId = int(self.classListWidget.item(row,1).text())
                subId = int(self.scoreSubTreeWidget.currentItem().whatsThis(0))
                params = []
                self.deleteScoreById(scoreId) #기존 스코어 삭제 후 
                params.append(subId)
                params.append(stdId)
                params.append(score)
                params.append(asses)
                self.saveScore(params) # 다시 재 저장.
            else: #기존에 스코어 존재 X
                if(self.classListWidget.item(row,2).whatsThis() == "" and self.classListWidget.item(row,2).text() != ""):
                    asses = self.classListWidget.item(row,3).text()
                    score = int(self.classListWidget.item(row,2).text())
                    stdId = int(self.classListWidget.item(row,1).text())
                    subId = int(self.scoreSubTreeWidget.currentItem().whatsThis(0))
                    params = []
                    params.append(subId)
                    params.append(stdId)
                    params.append(score)
                    params.append(asses)
                    self.saveScore(params) # 새로 저장.
        QMessageBox.about(self, "결과", "점수 저장 성공.")

    def returnScore(self, subId, stdId):
        conn = sqlite3.connect("studentManager.db")
        sql = "SELECT id, score, asses FROM Score WHERE subId = ? and stdId = ?"
        try:
            with conn:
                c = conn.cursor()
                return c.execute(sql, (int(subId), int(stdId),)).fetchone()
        except sqlite3.IntegrityError:
            print("점수 조회 오류")
            QMessageBox.about(self, "결과", "점수 조회 실패.")

    #점수, 평가 리스트 보여주는 함수
    def showScoreList(self):
        scoreInfo = []
        subId = int(self.scoreSubTreeWidget.currentItem().whatsThis(0))
        stdId = []
        assesments = []
        for row in range(0, self.classListWidget.rowCount()):
            stdId.append(int(self.classListWidget.item(row,1).text()))
        
        for row in range(0, self.classListWidget.rowCount()):
            if(self.returnScore(int(subId), int(stdId[row])) is not None):
                scoreInfo.append(self.returnScore(int(subId), int(stdId[row])))
            else:
                scoreInfo.append(["","",""])
        
        for row in range(0, self.classListWidget.rowCount()):
            asses = []
            if(scoreInfo[row][2] is not ""):
                asses.append(scoreInfo[row][2])
            else:
                asses.append("")
            assesments.append(asses)

        for row in range(0, self.classListWidget.rowCount()):
            if(scoreInfo[row][1] != ""):
                self.classListWidget.setItem(row, 2, QTableWidgetItem(str(scoreInfo[row][1]))) #점수입력
                self.classListWidget.item(row,2).setWhatsThis(str(scoreInfo[row][0])) #점수 id 속성
            else:
                self.classListWidget.setItem(row, 2, QTableWidgetItem(""))

            if(assesments[row][0] != ""):
                self.classListWidget.setItem(row, 3, QTableWidgetItem(str(assesments[row][0]))) #평가입력
                # self.classListWidget.item(row,3).setWhatsThis(str(scoreInfo[row][2])) #평가 id 속성
            else:
                self.classListWidget.setItem(row, 3, QTableWidgetItem(""))
            
    #점수 등급별 랜덤 평가 생성 함수 (선택)
    def insertIndiRandomAssesment(self):
        if(self.scoreSubTreeWidget.currentItem() is None):
            QMessageBox.about(self, "오류", "과목을 선택해주세요.")
            return
        clickedSub = self.scoreSubTreeWidget.currentItem().text(0)
        grdAList = []
        grdBList = []
        grdCList = []
        subId = self.returnSubIdBySubName(clickedSub)
        grdStandard = self.returnAssementStandardBySubId(subId)
        assementList = self.returnAssesmentBySubId(subId)
        for i in range(0, len(assementList)):
            if(assementList[i][1] == "A"):
                p = []
                p.append(assementList[i][0])
                p.append(assementList[i][2])
                grdAList.append(p)
            elif(assementList[i][1] == "B"):
                p = []
                p.append(assementList[i][0])
                p.append(assementList[i][2])
                grdBList.append(p)
            elif(assementList[i][1] == "C"):
                p = []
                p.append(assementList[i][0])
                p.append(assementList[i][2])
                grdCList.append(p)
        
        items = []
        items = self.classListWidget.selectedItems()
 
        for i in range(0, len(items)):
            if(self.classListWidget.selectedItems()[i].column() is 2):
                row = self.classListWidget.selectedItems()[i].row()
                score = self.classListWidget.item(row,2).text()
                if(score == ""):
                    self.classListWidget.setItem(row,3,QTableWidgetItem(""))
                else:
                    for j in range(0, len(grdStandard)):
                        # if(self.classListWidget.item(row,2) and self.classListWidget.item(row,2).text() != ""):
                        score = int(score)
                        if(grdStandard[j][1] < score and score <= grdStandard[j][2]):
                            if(grdStandard[j][0] == "A"):
                                randomIndex = random.randint(0, len(grdAList)-1)
                                self.classListWidget.setItem(row,3,QTableWidgetItem(grdAList[randomIndex][1]))
                                # self.classListWidget.item(row,3).setWhatsThis(str(grdAList[randomIndex][0]))
                            elif(grdStandard[j][0] == "B"):
                                randomIndex = random.randint(0, len(grdBList)-1)
                                self.classListWidget.setItem(row,3,QTableWidgetItem(grdBList[randomIndex][1]))
                                # self.classListWidget.item(row,3).setWhatsThis(str(grdBList[randomIndex][0]))
                            elif(grdStandard[j][0] == "C"):
                                randomIndex = random.randint(0, len(grdCList)-1)
                                self.classListWidget.setItem(row,3,QTableWidgetItem(grdCList[randomIndex][1]))
                                # self.classListWidget.item(row,3).setWhatsThis(str(grdCList[randomIndex][0]))

            

    #점수 등급별 랜덤 평가 생성 함수 (전체)
    def insertRandomAssesment(self):
        if(self.scoreSubTreeWidget.currentItem() is None):
            QMessageBox.about(self, "오류", "과목을 선택해주세요.")
            return
        grdAList = []
        grdBList = []
        grdCList = []
        subId = int(self.scoreSubTreeWidget.currentItem().whatsThis(0))
        grdStandard = self.returnAssementStandardBySubId(subId)
        assementList = self.returnAssesmentBySubId(subId)
        for i in range(0, len(assementList)):
            if(assementList[i][1] == "A"):
                p = []
                p.append(assementList[i][0])
                p.append(assementList[i][2])
                grdAList.append(p)
            elif(assementList[i][1] == "B"):
                p = []
                p.append(assementList[i][0])
                p.append(assementList[i][2])
                grdBList.append(p)
            elif(assementList[i][1] == "C"):
                p = []
                p.append(assementList[i][0])
                p.append(assementList[i][2])
                grdCList.append(p)

        for row in range(0, self.classListWidget.rowCount()):
            score = self.classListWidget.item(row,2).text()
            if(score == ""):
                self.classListWidget.setItem(row,3,QTableWidgetItem(""))
            else:
                for j in range(0, len(grdStandard)):
                    # if(self.classListWidget.item(row,2) and self.classListWidget.item(row,2).text() != ""):
                    score = int(score)
                    if(grdStandard[j][1] < score and score <= grdStandard[j][2]):
                        if(grdStandard[j][0] == "A"):
                            randomIndex = random.randint(0, len(grdAList)-1)
                            self.classListWidget.setItem(row,3,QTableWidgetItem(grdAList[randomIndex][1]))
                            # self.classListWidget.item(row,3).setWhatsThis(str(grdAList[randomIndex][0]))
                        elif(grdStandard[j][0] == "B"):
                            randomIndex = random.randint(0, len(grdBList)-1)
                            self.classListWidget.setItem(row,3,QTableWidgetItem(grdBList[randomIndex][1]))
                            # self.classListWidget.item(row,3).setWhatsThis(str(grdBList[randomIndex][0]))
                        elif(grdStandard[j][0] == "C"):
                            randomIndex = random.randint(0, len(grdCList)-1)
                            self.classListWidget.setItem(row,3,QTableWidgetItem(grdCList[randomIndex][1]))
                            # self.classListWidget.item(row,3).setWhatsThis(str(grdCList[randomIndex][0]))
    
    #학급 리스트 출력 함수
    def insertClassComboBox(self, combobox):
        classes = self.returnClassList()
        for i in range(0, len(classes)):
            combobox.addItem(str(classes[i][0])+"학년 "+str(classes[i][1])+"반")

    #선택 학급 조회 함수    
    def showClassMemberList(self):     
        grade = int(self.classList.currentText()[0])
        classes = int(self.classList.currentText()[4])
        members = self.returnClassMemberList(grade, classes)
        headers = ["이름", "학번", "점수", "평가"]

        self.classListWidget.setRowCount(len(members))
        self.classListWidget.setColumnCount(4)
        self.classListWidget.setHorizontalHeaderLabels(headers)

        for i in range(0, len(members)):
            name = QTableWidgetItem(str(members[i][0]))
            stdId = QTableWidgetItem(str(members[i][1]))
            self.classListWidget.setItem(i, 0, name)
            self.classListWidget.setItem(i, 1, stdId)
        
        self.showScoreList()
        

    ##############과목관리###########################

    def addNewSubjectItem(self):
        QTreeWidgetItem(self.subTreeWidget, ["새과목"])

    def addChildSub(self):
        parentItem = self.subTreeWidget.currentItem()
        childItem = QTreeWidgetItem(parentItem)
        childItem.setText(0, "새항목")
    
    #선택 과목 삭제 함수
    def deleteSubById(self, subId):
        conn = sqlite3.connect("studentManager.db")
        sql1 = "delete from Subject where parentId = ?"
        sql2 = "delete from Subject where id = ?"
        sql3 = "delete from Assesment where subId = ?"
        try:
            with conn:
                c = conn.cursor()
                c.execute(sql1, (subId,))
                c.execute(sql2, (subId,))
                c.execute(sql3, (subId,))
                return True  

        except sqlite3.IntegrityError:
            print("과목 삭제 문제 발생")
            return False

    def delSub(self):
        clickedItem = self.subTreeWidget.currentItem()
        if(clickedItem is None):
            QMessageBox.about(self, "결과", "삭제할 과목을 선택해주세요.")
            return
        if(clickedItem.whatsThis(0) == ''):
            self.subTreeWidget.removeItemWidget(clickedItem,0)
            self.showSub(self.subTreeWidget)
        else:
            clickedSubId = int(clickedItem.whatsThis(0))
            if(self.deleteSubById(clickedSubId)):
                QMessageBox.about(self, "결과", "삭제 성공")   
            else:
                QMessageBox.about(self, "결과", "삭제 실패")
        self.showSub(self.subTreeWidget)


    #과목 리스트에서 과목 선택 조회 하면 과목 세부 내용 조회 함수
    def searchSub(self):
        conn = sqlite3.connect("studentManager.db")
        subTreeWidget = self.subTreeWidget
        clickedSubId = subTreeWidget.currentItem().whatsThis(0)
        if(clickedSubId == ''):
            clickedSubId = -1
        sql2 = "select greater, less from Assesment where subId = ? and grade = ?"
        sql3 = "select content from Assesment where subId = ? and grade = ?"
        try:
            with conn:
                c = conn.cursor()
                subId = int(clickedSubId)
                grdARng = []
                grdBRng = []
                grdCRng = []
                if(subId):

                    #조회된 과목 등급 점수표 조회 및 출력
                    grdARng = c.execute(sql2, (subId, "A")).fetchall()
                    grdBRng = c.execute(sql2, (subId, "B")).fetchall()
                    grdCRng = c.execute(sql2, (subId, "C")).fetchall()

                    if(len(grdARng) != 0):
                        self.grdAEdit1.setText(str(grdARng[0][0]))
                        self.grdAEdit2.setText(str(grdARng[0][1]))
                    else:
                        self.grdAEdit1.setText(str("없음"))
                        self.grdAEdit2.setText(str("없음"))
                    if(len(grdBRng) != 0):
                        self.grdBEdit1.setText(str(grdBRng[0][0]))
                        self.grdBEdit2.setText(str(grdBRng[0][1]))
                    else:
                        self.grdBEdit1.setText(str("없음"))
                        self.grdBEdit2.setText(str("없음"))
                    if(len(grdCRng) != 0):
                        self.grdCEdit1.setText(str(grdCRng[0][0]))
                        self.grdCEdit2.setText(str(grdCRng[0][1]))
                    else:
                        self.grdCEdit1.setText(str("없음"))
                        self.grdCEdit2.setText(str("없음"))

                    #조회된 과목 이름 및 평가 출력
                    conn.row_factory = lambda cursor, row: row[0]
                    c = conn.cursor()
                    self.subTitleEdit.setText(self.subTreeWidget.currentItem().text(0))
                    
                    grdAAse = []
                    grdBAse = []
                    grdCAse = []

                    grdAAse = c.execute(sql3, (subId, "A")).fetchall()
                    grdBAse = c.execute(sql3, (subId, "B")).fetchall()
                    grdCAse = c.execute(sql3, (subId, "C")).fetchall()

                    self.grdAAseList.setRowCount(len(grdAAse))
                    self.grdAAseList.setColumnCount(1)
                    self.grdBAseList.setRowCount(len(grdBAse))
                    self.grdBAseList.setColumnCount(1)
                    self.grdCAseList.setRowCount(len(grdCAse))
                    self.grdCAseList.setColumnCount(1)

                    for i in range(0, len(grdAAse)):
                        self.grdAAseList.setItem(i, 0, QTableWidgetItem(grdAAse[i]))
                    for i in range(0, len(grdBAse)):
                        self.grdBAseList.setItem(i, 0, QTableWidgetItem(grdBAse[i]))
                    for i in range(0, len(grdCAse)):
                        self.grdCAseList.setItem(i, 0, QTableWidgetItem(grdCAse[i]))

        except sqlite3.IntegrityError:
            print("과목조회 문제 발생")


    #db 에서 과목 리스트 가져와서 보여주는 함수
    def returnSubList(self):
        conn = sqlite3.connect("studentManager.db")
        # conn.row_factory = lambda cursor, row: row[0]
        sql = "select id, subName, parentId from Subject"
        subList = []

        try:    
            with conn:
                c = conn.cursor()
                subList = c.execute(sql).fetchall()
                return subList
        except sqlite3.IntegrityError:
            print("문제 발생")

    def showSub(self, treeWidget):
        subList = self.returnSubList()
        subTreeWidget = treeWidget
        subTreeWidget.clear()
        subTreeWidget.setColumnCount(1)
        subTreeWidget.setHeaderLabels(["과목"])

        for i in range (0, len(subList)):
            if(subList[i][2] is None):
                subId = subList[i][0]
                subName = subList[i][1]
                parentItem = QTreeWidgetItem(subTreeWidget, [subName])
                parentItem.setWhatsThis(0,str(subId))

        for i in range(0, len(subList)):
            if(subList[i][2] is not None):
                subId = subList[i][0]
                subName = self.returnSubNameBySubId(subList[i][2])
                childName = subList[i][1]
                parentId = subList[i][2]
                it = QTreeWidgetItemIterator(subTreeWidget)
                while it.value():
                    if it.value() is not None and int(it.value().whatsThis(0)) == int(parentId):
                        childItem = QTreeWidgetItem(it.value())
                        childItem.setWhatsThis(0,str(subId))
                        childItem.setText(0,childName)
                    it += 1
                # parentItem = subTreeWidget.findItems(subName, QtCore.Qt.MatchExactly, column=0)
                # childItem = QTreeWidgetItem(parentItem[0])
                # childItem.setWhatsThis(0,str(subId))
                # childItem.setText(0,childName)


    #과목, 평가 등급 점수, 평가 내용 db 저장 함수
    def saveSub(self):
        subTitleEdit = self.subTreeWidget.currentItem().text(0)
        grdAEdit1 = self.grdAEdit1.text()
        grdAEdit2 = self.grdAEdit2.text()
        grdBEdit1 = self.grdBEdit1.text()
        grdBEdit2 = self.grdBEdit2.text()
        grdCEdit1 = self.grdCEdit1.text()
        grdCEdit2 = self.grdCEdit2.text()

        grdAList = []
        grdBList = []
        grdCList = []

        for i in range (0,self.grdAAseList.rowCount()):
            for j in range (0,self.grdAAseList.columnCount()):
                grdAList.append(self.grdAAseList.item(i,j).text())

        for i in range (0,self.grdBAseList.rowCount()):
            for j in range (0,self.grdBAseList.columnCount()):
                grdBList.append(self.grdBAseList.item(i,j).text())

        for i in range (0,self.grdCAseList.rowCount()):
            for j in range (0,self.grdCAseList.columnCount()):
                grdCList.append(self.grdCAseList.item(i,j).text())

        conn = sqlite3.connect("studentManager.db")

        #기존 과목 없으면 새로 생성
        try:    
            with conn:
                subName = self.subTitleEdit.text()
                subId = self.subTreeWidget.currentItem().whatsThis(0)
                c = conn.cursor()

                if(not subName):
                    QMessageBox.about(self, "오류", "과목을 입력하세요")
                else:
                    if(subId == ''): # 기존에 없던 과목인 경우
                        if(self.subTreeWidget.currentItem().parent()): #하위 노드일 경우
                            parent = self.subTreeWidget.currentItem().parent()
                            parentName = parent.text(0)
                            print("부모노드: ",parentName)
                            parentId = parent.whatsThis(0)
                            sql = "insert into Subject(subName, parentId) values (?, ?)"
                            c.execute(sql, (subName, int(parentId)))

                            sql2 = "select id from Subject where subName = ? and parentId = ?"
                            childSubId = c.execute(sql2, (subName, int(parentId))).fetchone()[0]
                            # print(childSubId)
                            c.execute("delete from Assesment where subId = ?", (int(childSubId),))
                            sql = "insert into Assesment(subId, grade, content, greater, less) values (?,?,?,?,?)"
                            for i in range(0, len(grdAList)):
                                c.execute(sql, (int(childSubId), "A", grdAList[i], int(grdAEdit1), int(grdAEdit2)))
                            
                            for i in range(0, len(grdBList)):
                                c.execute(sql, (int(childSubId), "B", grdBList[i], int(grdBEdit1), int(grdBEdit2)))

                            for i in range(0, len(grdCList)):
                                c.execute(sql, (int(childSubId), "C", grdCList[i], int(grdCEdit1), int(grdCEdit2)))

                            
                        else: #부모 노드일 경우
                            print(c.execute("insert into Subject(subName) values (?)", (subName,)))
                    else: # 기존에 있던 과목인 경우
                        
                        #기존 과목을 참조하여 평가지를 새로 생성 및 수정 (먼저 다 삭제했다가 다시 새로 생성)
                        # subId = c.execute("select id from Subject where subName=?", (subName,)).fetchone()[0]
                        c.execute("delete from Assesment where subId = ?", (int(subId),))
                        sql = "insert into Assesment(subId, grade, content, greater, less) values (?,?,?,?,?)"
                        for i in range(0, len(grdAList)):
                            c.execute(sql, (int(subId), "A", grdAList[i], int(grdAEdit1), int(grdAEdit2)))
                        
                        for i in range(0, len(grdBList)):
                            c.execute(sql, (int(subId), "B", grdBList[i], int(grdBEdit1), int(grdBEdit2)))

                        for i in range(0, len(grdCList)):
                            c.execute(sql, (int(subId), "C", grdCList[i], int(grdCEdit1), int(grdCEdit2)))

                    QMessageBox.about(self, "결과", "성공")
        except sqlite3.IntegrityError:
            print("문제 발생")
            QMessageBox.about(self, "오류", "실패")

        self.showSub(self.subTreeWidget)
        
    # 평가 내용 수정 함수
    def modAse(self):
        focusedTab = self.grdAseWidget.currentIndex()
        if(focusedTab == 0):
            widget = self.grdAAseList
            content = self.grdAseEdit.toPlainText()
            widget.setItem(widget.currentRow(), widget.currentColumn(), QTableWidgetItem(content))

        
        elif(focusedTab == 1):
            widget = self.grdBAseList
            content = self.grdAseEdit.toPlainText()
            widget.setItem(widget.currentRow(), widget.currentColumn(), QTableWidgetItem(content))

        elif(focusedTab == 2):
            widget = self.grdCAseList
            content = self.grdAseEdit.toPlainText()
            widget.setItem(widget.currentRow(), widget.currentColumn(), QTableWidgetItem(content))

    #평가 내용 선택하면 편집기에 해당 내용 보여줌
    def activateEdit(self):
        focusedTab = self.grdAseWidget.currentIndex()
        if(focusedTab == 0):
            widget = self.grdAAseList
            self.grdAseEdit.setPlainText(widget.item(widget.currentRow(), widget.currentColumn()).text())

        
        elif(focusedTab == 1):
            widget = self.grdBAseList
            self.grdAseEdit.setPlainText(widget.item(widget.currentRow(), widget.currentColumn()).text())

        elif(focusedTab == 2):
            widget = self.grdCAseList
            self.grdAseEdit.setPlainText(widget.item(widget.currentRow(), widget.currentColumn()).text())

    #평가 내용 항목 지우는 함수
    def delAse(self):
        focusedTab = self.grdAseWidget.currentIndex()
        
        if(focusedTab == 0):
            row = self.grdAAseList.currentRow()
            col = self.grdAAseList.currentColumn()
            self.grdAAseList.removeRow(row)
        
        elif(focusedTab == 1):
            row = self.grdBAseList.currentRow()
            col = self.grdBAseList.currentColumn()
            self.grdBAseList.removeRow(row)

        
        elif(focusedTab == 2):
            row = self.grdCAseList.currentRow()
            col = self.grdCAseList.currentColumn()
            self.grdCAseList.removeRow(row)

    #평가 내용 추가 함수
    def addAse(self):
        focusedTab = self.grdAseWidget.currentIndex()
        content = self.grdAseEdit.toPlainText()
        item = QTableWidgetItem(content)
                
        if(focusedTab == 0):
            currentRowCnt = self.grdAAseList.rowCount()
            if(currentRowCnt == 0):
                self.grdAAseList.insertRow(currentRowCnt)
                self.grdAAseList.setItem(currentRowCnt, 0, item)
            else:
                self.grdAAseList.insertRow(currentRowCnt)
                self.grdAAseList.setItem(currentRowCnt, 0, item)
        elif(focusedTab == 1):
            currentRowCnt = self.grdBAseList.rowCount()
            if(currentRowCnt == 0):
                self.grdBAseList.insertRow(currentRowCnt)
                self.grdBAseList.setItem(currentRowCnt, 0, item)
            else:
                self.grdBAseList.insertRow(currentRowCnt)
                self.grdBAseList.setItem(currentRowCnt, 0, item)
        
        elif(focusedTab == 2):
            currentRowCnt = self.grdCAseList.rowCount()
            if(currentRowCnt == 0):
                self.grdCAseList.insertRow(currentRowCnt)
                self.grdCAseList.setItem(currentRowCnt, 0, item)
            else:
                self.grdCAseList.insertRow(currentRowCnt)
                self.grdCAseList.setItem(currentRowCnt, 0, item)
        
    ##############학급추가###########################
    
    #엑셀로 불러온 학급 구성원 db 에 업로드 
    def saveStudent(self, name, grade, classes):
        conn = sqlite3.connect("studentManager.db")
        sql = "insert into Student(name, grade, class) values (?,?,?)"
        try:
            with conn:
                c = conn.cursor()
                c.execute(sql, (name, grade, classes))
                return True  

        except sqlite3.IntegrityError:
            print("과목 저장 문제 발생")
            QMessageBox.about(self, "결과", "학생 저장 실패.")
            return False


    def uploadCls(self):

        rowCnt = self.stdListWidget.rowCount()
        columnCnt = self.stdListWidget.columnCount()

        for col in range(0, columnCnt):
            header = self.stdListWidget.horizontalHeaderItem(col).text()
            grade = header[0]
            classes = header[2]
            for row in range(0, rowCnt):
                name = self.stdListWidget.item(row,col).text()
                if(name != "" or name is not None):
                    self.saveStudent(name, grade, classes)
                    
        QMessageBox.about(self, "결과", "학생 저장 완료.")
        



    #학급 구성원 엑셀 파일로 불러와서 리스트로 보여줌
    def uploadFile(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file', './')
        if(fname[0]):    
            wb = load_workbook(filename = fname[0]) 
            ws = wb["Sheet1"]
            max_row = ws.max_row
            max_col = ws.max_column

            widget = self.stdListWidget
            widget.setRowCount(max_row-1)
            widget.setColumnCount(max_col)
            headers = []
            for i in range(1, max_col+1):
                headers.append(str(ws.cell(row=1, column= i).value))
            widget.setHorizontalHeaderLabels(headers)

            for i in range(2, max_row+1):
                for j in range(1, max_col+1):
                    cellValue = str(ws.cell(row=i, column= j).value)
                    item = QTableWidgetItem(cellValue)
                    widget.setItem(i-2,j-1,item)

if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #WindowClass의 인스턴스 생성
    myWindow = WindowClass() 

    #프로그램 화면을 보여주는 코드
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()